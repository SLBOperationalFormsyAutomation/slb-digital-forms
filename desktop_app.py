import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import requests
import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
from PIL import Image
import base64
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as RLImage, Spacer
from reportlab.lib.pagesizes import landscape
from reportlab.lib.units import inch

# Configuración de Supabase
SUPABASE_URL = "https://jfhsgobubmmedsbtenay.supabase.co"
SUPABASE_ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImpmaHNnb2J1Ym1tZWRzYnRlbmF5Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzQ4OTM2NjgsImV4cCI6MjA5MDQ2OTY2OH0.QGs3_rYsX7e4uzUMJhtLkL47-rcdhMtZWWElsLptmrI"  # Clave real

def formatear_fecha(fecha_str):
    """Convierte fecha ISO a formato DD/MM/YYYY HH:MM:SS"""
    try:
        if 'T' in fecha_str:
            dt = datetime.fromisoformat(fecha_str.replace('Z', '+00:00'))
            return dt.strftime('%d/%m/%Y %H:%M:%S')
        return fecha_str
    except:
        return fecha_str

def descargar_reporte(formato):
    try:
        # URL para obtener todos los registros
        url = f"{SUPABASE_URL}/rest/v1/registros"
        headers = {
            "apikey": SUPABASE_ANON_KEY,
            "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
        }
        params = {
            "select": "*"  # Obtener todos los campos
        }
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        data = response.json()

        if not data:
            messagebox.showinfo("Información", "No hay registros para descargar.")
            return

        # Pedir ubicación para guardar
        if formato == "Excel":
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Guardar Reporte Excel",
                initialfile=f"reporte_visitantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            if not file_path:
                return
            # Generar Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Registros"

            # Headers - Excluir created_at
            all_headers = list(data[0].keys())
            headers = [h for h in all_headers if h != 'created_at']
            
            # Estilo de encabezados (Azul SLB)
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            ws.row_dimensions[1].height = 30

            # Datos
            for row_num, reg in enumerate(data, 2):
                row_height = 20
                for col_num, key in enumerate(headers, 1):
                    if key == "firma" and reg.get(key):
                        # Insertar imagen de firma
                        try:
                            firma_base64 = reg[key]
                            if isinstance(firma_base64, str) and ',' in firma_base64:
                                firma_data = base64.b64decode(firma_base64.split(',')[1])
                            elif isinstance(firma_base64, str):
                                firma_data = base64.b64decode(firma_base64)
                            else:
                                raise ValueError("Formato de firma inválido")
                            img = Image.open(BytesIO(firma_data))
                            img_resized = img.resize((80, 60))
                            img_temp = BytesIO()
                            img_resized.save(img_temp, format='PNG')
                            img_temp.seek(0)
                            xl_img = XLImage(img_temp)
                            ws.add_image(xl_img, f'{chr(64 + col_num)}{row_num}')
                            row_height = 65
                        except Exception as e:
                            ws.cell(row=row_num, column=col_num, value="Error en firma")
                    elif key == "fechaRegistro" and reg.get(key):
                        ws.cell(row=row_num, column=col_num, value=formatear_fecha(reg[key]))
                    else:
                        ws.cell(row=row_num, column=col_num, value=reg.get(key, ""))
                
                ws.row_dimensions[row_num].height = row_height
            
            # Ajustar ancho de columnas
            for col_num, key in enumerate(headers, 1):
                if key == "firma":
                    ws.column_dimensions[chr(64 + col_num)].width = 15
                else:
                    ws.column_dimensions[chr(64 + col_num)].width = 18

            wb.save(file_path)

        elif formato == "PDF":
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Guardar Reporte PDF",
                initialfile=f"reporte_visitantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
            if not file_path:
                return
            # Generar PDF
            doc = SimpleDocTemplate(file_path, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.3*inch, rightMargin=0.3*inch)
            elements = []

            # Datos para tabla - Excluir created_at
            all_keys = list(data[0].keys())
            headers = [h for h in all_keys if h != 'created_at']
            table_data = [headers]  # Headers
            
            for reg in data:
                row = []
                for key in headers:
                    if key == "firma" and reg.get(key):
                        # Para PDF, insertar imagen
                        try:
                            firma_base64 = reg[key]
                            if ',' in firma_base64:
                                firma_data = base64.b64decode(firma_base64.split(',')[1])
                            else:
                                firma_data = base64.b64decode(firma_base64)
                            img = Image.open(BytesIO(firma_data))
                            img_path = f"temp_firma_pdf_{len(table_data)}.png"
                            img.save(img_path)
                            rl_img = RLImage(img_path, width=28, height=18)
                            row.append(rl_img)
                        except:
                            row.append("Firma")
                    elif key == "fechaRegistro" and reg.get(key):
                        row.append(formatear_fecha(reg[key]))
                    else:
                        row.append(str(reg.get(key, ""))[:30])  # Limitar texto a 30 caracteres
                table_data.append(row)

            # Crear tabla con ancho optimizado para mostrar todas las columnas
            # Calcular ancho dinámico basado en número de columnas
            num_cols = len(headers)
            available_width = 10.5 * inch  # Ancho disponible en landscape
            col_width = available_width / num_cols
            col_widths = [col_width] * num_cols
            
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
                ('TOPPADDING', (0, 0), (-1, 0), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 1), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(table)
            doc.build(elements)
            # Limpiar archivos temp
            for i in range(len(table_data)):
                temp_path = f"temp_firma_pdf_{i}.png"
                if os.path.exists(temp_path):
                    os.remove(temp_path)

        messagebox.showinfo("Éxito", f"Reporte descargado exitosamente en {file_path}")

    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 401:
            messagebox.showerror("Error", "Clave API incorrecta. Verifica la configuración.")
        else:
            messagebox.showerror("Error", f"Error HTTP: {e.response.status_code}")
    except Exception as e:
        messagebox.showerror("Error", f"Error inesperado: {str(e)}")

# Crear ventana principal
root = tk.Tk()
root.title("Descargador de Reportes - Registro de Visitantes SLB")
root.geometry("500x450")
root.configure(bg="#003366")

# Estilo
style = ttk.Style()
style.configure("TLabel", background="#003366", foreground="white", font=("Arial", 10))
style.configure("TButton", font=("Arial", 10, "bold"))
style.configure("TCombobox", font=("Arial", 10))

# Frame principal
main_frame = tk.Frame(root, bg="#003366")
main_frame.pack(pady=20, padx=20, fill="both", expand=True)

# Logo
logo_label = tk.Label(main_frame, text="SLB", font=("Arial", 24, "bold"), fg="white", bg="#003366")
logo_label.pack(pady=10)

subtitle_label = tk.Label(main_frame, text="Sistema de Registro de Visitantes", font=("Arial", 12), fg="white", bg="#003366")
subtitle_label.pack(pady=5)

# Selector de formato
format_frame = tk.Frame(main_frame, bg="#003366")
format_frame.pack(pady=10)
format_label = tk.Label(format_frame, text="Formato del reporte:", fg="white", bg="#003366", font=("Arial", 10))
format_label.pack(side="left")
format_combo = ttk.Combobox(format_frame, values=["Excel", "PDF"], state="readonly")
format_combo.current(0)  # Default Excel
format_combo.pack(side="left", padx=10)

# Botón
button = tk.Button(main_frame, text="Descargar Reporte", command=lambda: descargar_reporte(format_combo.get()), 
                   height=2, width=25, bg="white", fg="#003366", font=("Arial", 12, "bold"))
button.pack(pady=20)

# Instrucciones
instructions = tk.Label(main_frame, text="Selecciona el formato y haz clic para descargar.\nExcel y PDF incluyen encabezados con colores SLB.", 
                        fg="white", bg="#003366", font=("Arial", 9), wraplength=400, justify="center")
instructions.pack(pady=10)

# Ejecutar la aplicación
root.mainloop()