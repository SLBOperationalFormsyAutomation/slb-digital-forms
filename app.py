from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timedelta, timezone
import base64
from io import BytesIO
import os
import requests
from PIL import Image

app = Flask(__name__)
CORS(app)

SUPABASE_URL = os.getenv("SUPABASE_URL", "https://jfhsgobubmmedsbtenay.supabase.co")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY", "TU_API_KEY_AQUI")

# 🔹 GUARDAR REGISTRO
def guardar_registro(data):
    url = f"{SUPABASE_URL}/rest/v1/registros"

    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }

    # 🇨🇴 Hora Colombia
    colombia = timezone(timedelta(hours=-5))
    fecha_colombia = datetime.now(colombia).isoformat()

    payload = {
        "nombre": data["nombre"],
        "tipoDocumento": data["tipoDocumento"],
        "documento": data["documento"],
        "contacto": data["contacto"],
        "empresa": data["empresa"],
        "alcocheck": data["alcocheck"],
        "arl": data.get("arl",""),
        "eps": data.get("eps",""),
        "rh": data.get("rh",""),
        "alergias": data.get("alergias",""),
        "emergencia": data.get("emergencia",""),
        "visita": data.get("visita",""),
        "serial": data.get("serial",""),
        "laptopIngreso": data.get("laptopIngreso",""),
        "laptopSalida": data.get("laptopSalida",""),
        "aceptaDatos": data["aceptaDatos"],
        "firma": data["firma"],
        "fechaRegistro": fecha_colombia
    }

    r = requests.post(url, json=payload, headers=headers)
    r.raise_for_status()
    return r.json()

# 🔹 HOME
@app.route("/", methods=["GET"])
def home():
    return render_template("index.html")

# 🔹 STATUS
@app.route("/status", methods=["GET"])
def status():
    return jsonify({
        "status": "ok",
        "app": "visitor-registration",
        "time": datetime.now(timezone.utc).isoformat()
    })

# 🔹 OBTENER REGISTROS
@app.route("/registros", methods=["GET"])
def get_registros():
    url = f"{SUPABASE_URL}/rest/v1/registros"

    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {SUPABASE_ANON_KEY}"
    }

    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return jsonify(response.json())

# 🔹 REGISTRO
@app.route("/registro", methods=["POST"])
def registro():
    data = request.json

    if not data.get("aceptaDatos"):
        return jsonify({"error": "Debe aceptar los términos"}), 400

    try:
        guardar_registro(data)
        return jsonify({"ok": True})
    except Exception as e:
        app.logger.exception("Error en /registro")
        return jsonify({"error": str(e)}), 500

# 🔹 FORMATEAR FECHA
def formatear_fecha(fecha_str):
    try:
        if not fecha_str:
            return ""

        dt = datetime.fromisoformat(str(fecha_str).replace("Z", "+00:00"))

        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)

        colombia = timezone(timedelta(hours=-5))
        dt_col = dt.astimezone(colombia)

        return dt_col.strftime('%d/%m/%Y %H:%M:%S')

    except Exception as e:
        print("Error fecha:", e, fecha_str)
        return str(fecha_str)

# 🔹 EXPORTAR EXCEL
@app.route("/export_excel", methods=["GET"])
def export_excel():

    url = f"{SUPABASE_URL}/rest/v1/registros"

    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {SUPABASE_ANON_KEY}"
    }

    response = requests.get(url, headers=headers)
    response.raise_for_status()
    registros = response.json()

    if not registros:
        return jsonify({"error": "No hay datos"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    # 🔹 KEYS (orden real)
    keys = [
        "ID",
        "nombre",
        "tipoDocumento",
        "documento",
        "contacto",
        "empresa",
        "alcocheck",
        "arl",
        "eps",
        "rh",
        "alergias",
        "emergencia",
        "visita",
        "serial",
        "laptopIngreso",
        "laptopSalida",
        "aceptaDatos",
        "firma",
        "fechaRegistro"
    ]

    # 🔹 HEADERS BONITOS
    headers_excel = [
        "ID",
        "Nombre",
        "Tipo de Documento",
        "Documento",
        "Contacto",
        "Empresa",
        "Alcocheck",
        "ARL",
        "EPS",
        "RH",
        "Alergias",
        "Contacto de Emergencia",
        "A quien visita",
        "Serial Laptop",
        "Fecha de Ingreso Laptop",
        "Fecha de Salida Laptop",
        "Acepta tratamiento de Datos",
        "Firma",
        "Fecha de Registro Formulario"
    ]

    # 🔹 ESTILO HEADER
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(headers_excel, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 30

    # 🔹 DATA
    for i, reg in enumerate(registros, start=1):
        row_num = i + 1
        row_height = 20

        for col_num, key in enumerate(keys, 1):

            if key == "ID":
                ws.cell(row=row_num, column=col_num, value=i)

            elif key == "firma" and reg.get("firma"):
                try:
                    firma_base64 = reg["firma"]

                    if ',' in firma_base64:
                        firma_data = base64.b64decode(firma_base64.split(',')[1])
                    else:
                        firma_data = base64.b64decode(firma_base64)

                    img = Image.open(BytesIO(firma_data))
                    img = img.resize((80, 60))

                    img_bytes = BytesIO()
                    img.save(img_bytes, format="PNG")
                    img_bytes.seek(0)

                    xl_img = XLImage(img_bytes)
                    ws.add_image(xl_img, f'{chr(64 + col_num)}{row_num}')

                    row_height = 65

                except:
                    ws.cell(row=row_num, column=col_num, value="Error firma")

            elif key == "fechaRegistro":
                ws.cell(row=row_num, column=col_num, value=formatear_fecha(reg.get(key)))

            else:
                ws.cell(row=row_num, column=col_num, value=reg.get(key, ""))

        ws.row_dimensions[row_num].height = row_height

    # 🔹 AUTO WIDTH
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="registros_visitantes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    app.run(debug=True)